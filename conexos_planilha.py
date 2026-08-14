"""
conexos_planilha.py
--------------------
Conector REAL do CONEXOS, baseado na planilha exportada manualmente da
tela "Gerência de Processos" (Excel/CSV) — ponte rápida enquanto não
existe API oficial do CONEXOS.

Fluxo operacional (combinado com o Bernardo em 14/08):
  1. No CONEXOS, filtrar "Gerência de Processos" pelas filiais 4 e 5,
     todos os status, todos os processos em aberto, e exportar pra
     Excel.
  2. Enviar esse arquivo pro endpoint POST /admin/upload-conexos (ver
     app.py) — protegido por senha (UPLOAD_SECRET). Isso salva o
     arquivo no servidor, no caminho de CONEXOS_EXPORT_PATH.
  3. As rotinas de e-mail (followup_diario.py, followup_semanal.py)
     passam a ler os dados reais desse arquivo, em vez do mock.

Colunas confirmadas na planilha real exportada em 14/08/2026 (aba
única, cabeçalho na linha 1):
  Filial, Processo, Cód. Pessoa, Descrição da Pessoa, Ref. Externa,
  Mercadoria, Ref. Cliente, Conhec. Transp., Status do Processo,
  Previsão Carga Pronta, Data Carga Pronta, ETD, ETA,
  Valor FOB da Invoice, Total M. Neg., Solicitação de numerário,
  Exportador, Agente de carga, Descrição Despachante, Finalidade,
  Via de Transp., Incoterm, Origem, País de Origem, Destino,
  País Destino, Dias Transit Time, Qtd. De Containers,
  Número DI/DUIMP, Responsável, Situação do Processo,
  Histórico do Processo.

IMPORTANTE: essa planilha NÃO tem telefone do cliente nem CNPJ — então
buscar_processos_por_telefone() (usada pelo chat 1:1 do bot_engine)
não funciona com esse conector. Ele serve para as rotinas de e-mail
(resumo semanal e acompanhamento diário), que não precisam de telefone.

Campos que a planilha NÃO tem (ficam sempre None — nunca inventados):
  info_aduaneira.exigencia/pendencias/canal/desembaraco/conferencia,
  info_financeira (pendências, ICMS, demurrage não vêm num campo
  próprio — o que existe é "Solicitação de numerário", que é anexado
  como texto livre em info_financeira.pagamentos quando presente),
  info_entrega (nenhum campo de entrega/coleta vem nessa planilha).
"""

import os
from datetime import datetime, date

import openpyxl

CONEXOS_EXPORT_PATH = os.environ.get("CONEXOS_EXPORT_PATH", "dados_conexos.xlsx")

# Nomes exatamente como aparecem na coluna "Responsável" da planilha real.
NOMES_ESPERADOS_RESPONSAVEL = {"BERNARDO", "ABNERH", "SANTOSTHIAGO"}


def _valor(row_dict, nome_coluna):
    return row_dict.get(nome_coluna)


def _para_data_iso(valor):
    """Converte célula de data do Excel (datetime/date) pra 'YYYY-MM-DD'. None fica None."""
    if valor is None:
        return None
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")
    texto = str(valor).strip()
    return texto or None


def _para_texto(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _mapear_linha(row_dict: dict) -> dict:
    filial = _valor(row_dict, "Filial")
    codigo_empresa = str(int(filial)) if filial is not None else None

    processo_codigo = _para_texto(_valor(row_dict, "Ref. Externa")) or _para_texto(_valor(row_dict, "Processo"))
    status = _para_texto(_valor(row_dict, "Status do Processo"))

    numerario = _para_texto(_valor(row_dict, "Solicitação de numerário"))

    return {
        "processo": processo_codigo,
        "cliente": _para_texto(_valor(row_dict, "Descrição da Pessoa")),
        "cliente_cnpj": None,  # não vem nessa planilha
        "telefone_cliente": None,  # não vem nessa planilha (ver limitação no topo do arquivo)
        "responsavel_interno": _para_texto(_valor(row_dict, "Responsável")),
        "status_processo": status,
        "codigo_empresa": codigo_empresa,
        "referencia_cliente": _para_texto(_valor(row_dict, "Ref. Cliente")),
        "ultima_atualizacao": None,
        "info_processo": {
            "fornecedor": _para_texto(_valor(row_dict, "Exportador")),
            "origem": _para_texto(_valor(row_dict, "Origem")),
            "destino": _para_texto(_valor(row_dict, "Destino")),
            "modal": _para_texto(_valor(row_dict, "Via de Transp.")),
            "incoterm": _para_texto(_valor(row_dict, "Incoterm")),
            "tipo_operacao": _para_texto(_valor(row_dict, "Finalidade")),
            "data_abertura": None,
            "etapa_atual": status,
        },
        "info_carga": {
            "container": None,
            "bl": _para_texto(_valor(row_dict, "Conhec. Transp.")),
            "awb": None,
            "booking": None,
            "navio": None,
            "voo": None,
            "porto_aeroporto": None,
            "eta": _para_data_iso(_valor(row_dict, "ETA")),
            "etd": _para_data_iso(_valor(row_dict, "ETD")),
            "prontidao_carga_prevista": _para_data_iso(_valor(row_dict, "Previsão Carga Pronta")),
            "prontidao_carga_real": _para_data_iso(_valor(row_dict, "Data Carga Pronta")),
            "ata": None,
            "data_embarque": None,
            "data_chegada": None,
            "terminal": None,
            "transportadora": None,
            "agente_cargas": _para_texto(_valor(row_dict, "Agente de carga")),
        },
        "info_aduaneira": {
            "registro_di": _para_texto(_valor(row_dict, "Número DI/DUIMP")),
            "registro_duimp": None,
            "canal": None,
            "desembaraco": None,
            "conferencia": None,
            "exigencia": None,
            "pendencias": [],
            "liberacao": None,
            "situacao_rfb": None,
            "orgaos_anuentes": [],
            "licenciamento": None,
        },
        "info_financeira": {
            "pagamentos": numerario,
            "custos": None,
            "taxas": None,
            "icms": None,
            "afrmm": None,
            "armazenagem": None,
            "demurrage": None,
            "pendencias_financeiras": [],
        },
        "info_entrega": {
            "liberacao_carga": None,
            "nf": None,
            "transportadora_entrega": None,
            "agendamento": None,
            "coleta": None,
            "data_prevista_entrega": None,
            "tipo_previsao_entrega": None,
            "data_efetiva_entrega": None,
        },
        "historico": [],
    }


def _ler_planilha() -> list:
    if not os.path.exists(CONEXOS_EXPORT_PATH):
        raise NotImplementedError(
            f"Nenhuma planilha do CONEXOS encontrada em '{CONEXOS_EXPORT_PATH}'. "
            "Envie o export mais recente via POST /admin/upload-conexos."
        )

    wb = openpyxl.load_workbook(CONEXOS_EXPORT_PATH, data_only=True)
    ws = wb.active

    cabecalho = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    processos = []
    for r in range(2, ws.max_row + 1):
        valores = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in valores):
            continue
        row_dict = dict(zip(cabecalho, valores))
        processos.append(_mapear_linha(row_dict))
    return processos


def listar_todos() -> list:
    return _ler_planilha()


def buscar_processo_por_identificador(identificador: str):
    ident = (identificador or "").strip().upper().replace(" ", "")
    for p in _ler_planilha():
        if (p.get("processo") or "").strip().upper().replace(" ", "") == ident:
            return p
    return None


def buscar_processos_por_telefone(telefone: str):
    """Essa planilha não tem telefone do cliente — sempre retorna vazio.
    Use erp_mock ou conexos_connector (API) para o chat 1:1 do bot_engine."""
    return []
